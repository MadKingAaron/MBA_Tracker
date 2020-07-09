from openpyxl import *
import load_and_save_file_dialog
from openpyxl.styles import Font
from copy import copy
import course_values
import course_filters


class Spreadsheet:
    def __init__(self):
        self.fileName = load_and_save_file_dialog.getOpenDir()
        self.workbook = load_workbook(filename=self.fileName)

        self.setWorksheets()

        self.blankCellStyle = self.Style(self.main_worksheet.cell(row=222, column=1))

        self.clipboard = []

        self.fontClipboard = []

        self.table_startingRow = 3

        self.table_startingCol = 1

        self.table_endCol = 38

        self.table_endRow = 3

        self.findEndRow()

        self.findRangesForCohorts()

        self.copyMainWSTableToClipboard()

        self.saveDir = load_and_save_file_dialog.getSaveDir()

        pass

    def saveWorkbook(self):
        print("Saving...")
        self.workbook.save(filename=self.saveDir)
        print("Saved")

    def setWorksheets(self):
        worksheets = self.workbook.worksheets

        for i in range(0, len(worksheets)):
            worksheets[i] = worksheets[i].title

        start_found_index = worksheets.index("500 (All)")
        start_core_index = worksheets.index("510 (All)")

        foundation_courses_names = worksheets[start_found_index:start_core_index]

        core_courses_names = worksheets[start_core_index:]

        special_topics_names = []

        copy_core_course_names = core_courses_names.copy()

        for course in core_courses_names:
            if course[2] != '0':
                special_topics_names.append(course)
                copy_core_course_names.remove(course)

        core_courses_names = copy_core_course_names.copy()
        copy_core_course_names.clear()

        # Create dictionaries for each course with worksheet
        self.foundation_courses = {
            foundation_courses_names[i]: self.FoundationWorksheet(self, foundation_courses_names[i]) for i
            in range(0, len(foundation_courses_names))}

        self.core_courses = {core_courses_names[i]: self.CoreWorksheet(self, core_courses_names[i]) for i in
                             range(0, len(core_courses_names))}

        self.special_topics_courses = {special_topics_names[i]: self.STWorksheet(self, special_topics_names[i]) for i in
                                       range(0, len(special_topics_names))}

        self.main_worksheet = self.workbook[worksheets[0]]

    def copyMainWSTableToClipboard(self):
        rangeSelected = []
        fontRange = []
        # Loops through selected Rows
        for i in range(self.table_startingRow, self.table_endRow + 1, 1):
            # Appends the row to a RowSelected list
            rowSelected = []
            fontRow = []
            for j in range(self.table_startingCol, self.table_endCol + 1, 1):
                rowSelected.append(self.main_worksheet.cell(row=i, column=j).value)
                fontRow.append(self.Style(self.main_worksheet.cell(row=i, column=j)))
            # Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)
            fontRange.append(fontRow)
        self.clipboard = rangeSelected
        self.fontClipboard = fontRange

    def copyClipboardToAllWS(self):

        self.copyClipboardToFoundationWS()

        self.copyClipboardToCoreWS()

        self.copyClipboardToSpecialTopicsWS()

        pass

    def copyClipboardToFoundationWS(self):
        for course in self.foundation_courses:
            self.foundation_courses[course].clearRange()
            self.foundation_courses[course].pasteRange(self.table_startingCol, self.table_startingRow,
                                                       self.table_endCol, self.table_endRow)
            self.foundation_courses[course].calculateCounts()

    def copyClipboardToCoreWS(self):
        for course in self.core_courses:
            self.core_courses[course].clearRange()
            self.core_courses[course].pasteRange(self.table_startingCol, self.table_startingRow, self.table_endCol,
                                                 self.table_endRow)
            self.core_courses[course].calculateCounts()

    def copyClipboardToSpecialTopicsWS(self):
        for course in self.special_topics_courses:
            self.special_topics_courses[course].clearRange()
            self.special_topics_courses[course].pasteRange(self.table_startingCol, self.table_startingRow,
                                                           self.table_endCol, self.table_endRow)
            self.special_topics_courses[course].calculateCounts()

    def findEndRow(self):
        row = 3
        col = 4

        while self.main_worksheet.cell(row=row, column=col).value is not None:
            cell = self.main_worksheet.cell(row=row, column=col).value
            row += 1

        row -= 1

        self.table_endRow = row

    def findRangesForCohorts(self):
        self.findRangeForMBA()

        self.findRangeForBusCert()

        self.findRangeForDD()

    def findRangeForMBA(self):
        row = 3
        col = 4
        # Find start row for MBA
        while (self.main_worksheet.cell(row=row,
                                        column=col).value is not None) and "MBA" not in self.main_worksheet.cell(
            row=row, column=col).value:
            row += 1

        self.MBA_start_row = row

        # Find end row for MBA
        while (self.main_worksheet.cell(row=row, column=col).value is not None) and "MBA" in self.main_worksheet.cell(
                row=row, column=col).value:
            row += 1

        row -= 1

        self.MBA_end_row = row

    def findRangeForBusCert(self):
        row = 3
        col = 4
        # Find start row for Bus Cert
        while (self.main_worksheet.cell(row=row,
                                        column=col).value is not None) and "Bus Certificate" not in self.main_worksheet.cell(
            row=row, column=col).value:
            row += 1

        self.Bus_Cert_start_row = row

        # Find end row for MBA
        while (self.main_worksheet.cell(row=row,
                                        column=col).value is not None) and "Bus Certificate" in self.main_worksheet.cell(
            row=row, column=col).value:
            row += 1

        row -= 1

        self.Bus_Cert_end_row = row

    def findRangeForDD(self):
        row = 3
        col = 4
        # Find start row for DD
        cell = self.main_worksheet.cell(row=row, column=col)
        while (self.main_worksheet.cell(row=row,
                                        column=col).value is not None) and "DD" not in self.main_worksheet.cell(row=row,
                                                                                                                column=col).value:
            row += 1

        self.DD_start_row = row

        # Find end row for MBA
        while (self.main_worksheet.cell(row=row, column=col).value is not None) and (
                "DD" in self.main_worksheet.cell(row=row, column=col).value):
            row += 1

        row -= 1

        self.DD_end_row = row

    class Worksheet:
        def __init__(self, current_spreadsheet, current_worksheet_name):
            self.current_spreadsheet = current_spreadsheet
            self.current_worksheet = current_spreadsheet.workbook[current_worksheet_name]
            self.current_worksheet_name = current_worksheet_name
            self.filter_function = course_filters.course_filter_selector(self.current_worksheet_name)

        def copyRange(self, startRow, startCol, endRow, endCol):
            rangeSelected = []
            # Loops through selected Rows
            for i in range(startRow, endRow + 1, 1):
                # Appends the row to a RowSelected list
                rowSelected = []
                for j in range(startCol, endCol + 1, 1):
                    rowSelected.append(self.current_worksheet.cell(row=i, column=j).value)
                # Adds the RowSelected List and nests inside the rangeSelected
                rangeSelected.append(rowSelected)

            # Set range to clipboard
            self.current_spreadsheet.clipboard = rangeSelected

        def pasteRange(self, startCol, startRow, endCol, endRow):
            print(self.current_worksheet_name)
            countRow = 0
            countSheetRow = startRow

            for i in range(startRow, endRow + 1, 1):
                countCol = 0

                # Filter row
                if self.filter_function(self.current_spreadsheet.clipboard[countRow]) is True:
                    self.current_spreadsheet.clipboard[countRow][8] = course_filters.count_core_courses(
                        self.current_spreadsheet.clipboard[countRow])
                    self.current_spreadsheet.clipboard[countRow][7] = course_filters.count_found_courses(
                        self.current_spreadsheet.clipboard[countRow])
                    # print(self.current_spreadsheet.clipboard[countRow][1])
                    for j in range(startCol, endCol + 1, 1):
                        self.current_worksheet.cell(row=countSheetRow, column=j).value = \
                            self.current_spreadsheet.clipboard[countRow][
                                countCol]
                        self.current_spreadsheet.fontClipboard[countRow][countCol].copyToCell(
                            self.current_worksheet.cell(row=countSheetRow, column=j))

                        countCol += 1
                    countSheetRow += 1

                countRow += 1

        def clearRange(self, startCol=1, startRow=3, endCol=38):
            endRow = self.findEndRow()
            for i in range(startRow, endRow + 1, 1):
                for j in range(startCol, endCol + 1, 1):
                    self.current_worksheet.cell(row=i, column=j).value = None
                    self.clearStyle(self.current_worksheet.cell(row=i, column=j))

        def findEndRow(self, row=3, col=4):

            while self.current_worksheet.cell(row=row, column=col).value is not None:
                row += 1

            row -= 1
            return row

        def clearStyle(self, cell):
            self.current_spreadsheet.blankCellStyle.copyToCell(cell)

        def findRangeForMBA(self):
            row = 3
            col = 4
            # Find start row for MBA
            while (self.current_worksheet.cell(row=row,
                                               column=col).value is not None) and "MBA" not in self.current_worksheet.cell(
                row=row, column=col).value:
                row += 1

            # Check if in range
            if self.current_worksheet.cell(row=row, column=col).value is None:
                return -1, -1

            MBA_start_row = row

            # Find end row for MBA
            while (self.current_worksheet.cell(row=row,
                                               column=col).value is not None) and "MBA" in self.current_worksheet.cell(
                row=row, column=col).value:
                row += 1

            row -= 1

            MBA_end_row = row

            return MBA_start_row, MBA_end_row

        def findRangeForBusCert(self):
            row = 3
            col = 4
            # Find start row for Bus Cert
            while (self.current_worksheet.cell(row=row,
                                               column=col).value is not None) and "Bus Certificate" not in self.current_worksheet.cell(
                row=row, column=col).value:
                row += 1

            # Check if in range
            if self.current_worksheet.cell(row=row, column=col).value is None:
                return -1, -1

            Bus_Cert_start_row = row

            # Find end row for MBA
            while (self.current_worksheet.cell(row=row,
                                               column=col).value is not None) and "Bus Certificate" in self.current_worksheet.cell(
                row=row, column=col).value:
                row += 1

            row -= 1

            Bus_Cert_end_row = row

            return Bus_Cert_start_row, Bus_Cert_end_row

        def findRangeForDD(self):
            row = 3
            col = 4
            # Find start row for DD
            # cell = self.current_worksheet.cell(row=row, column=col)
            while (self.current_worksheet.cell(row=row,
                                               column=col).value is not None) and "DD" not in self.current_worksheet.cell(
                row=row, column=col).value:
                row += 1

            # Check if in range
            if self.current_worksheet.cell(row=row, column=col).value is None:
                return -1, -1

            DD_start_row = row

            # Find end row for MBA
            while (self.current_worksheet.cell(row=row, column=col).value is not None) and (
                    "DD" in self.current_worksheet.cell(row=row, column=col).value):
                row += 1

            row -= 1

            DD_end_row = row

            return DD_start_row, DD_end_row

        def setRangeForMBA(self):
            mbaRange = self.findRangeForMBA()

            if mbaRange[0] == -1 and mbaRange[1] == -1:
                MBA_subtotal = 0

            else:
                MBA_subtotal = "=SUBTOTAL(3, D" + str(mbaRange[0]) + ":D" + str(
                    mbaRange[1]) + ")"

            self.current_worksheet.cell(row=169, column=4).value = MBA_subtotal

        def setRangeForDD(self):
            ddRange = self.findRangeForDD()

            if ddRange[0] == -1 and ddRange[1] == -1:
                DD_subtotal = 0
            else:
                DD_subtotal = "=SUBTOTAL(3, D" + str(ddRange[0]) + ":D" + str(
                    ddRange[1]) + ")"

            self.current_worksheet.cell(row=170, column=4).value = DD_subtotal

        def setRangeForBusCert(self):
            busCertRange = self.findRangeForBusCert()

            if busCertRange[0] == -1 and busCertRange[1] == -1:
                bus_cert_subtotal = 0
            else:
                bus_cert_subtotal = "=SUBTOTAL(3, D" + str(busCertRange[0]) + ":D" + str(
                    busCertRange[1]) + ")"

            self.current_worksheet.cell(row=171, column=4).value = bus_cert_subtotal

    class CoreWorksheet(Worksheet):
        def __init__(self, current_spreadsheet, current_worksheet_name):
            super().__init__(current_spreadsheet, current_worksheet_name)

        def calculateCounts(self):
            self.setRangeForMBA()

            if "570" not in self.current_worksheet_name:
                self.setRangeForDD()

    class FoundationWorksheet(Worksheet):
        def __init__(self, current_spreadsheet, current_worksheet_name):
            super().__init__(current_spreadsheet, current_worksheet_name)

        def calculateCounts(self):
            self.setRangeForMBA()
            self.setRangeForDD()
            self.setRangeForBusCert()

    class STWorksheet(Worksheet):
        def __init__(self, current_spreadsheet, current_worksheet_name):
            super().__init__(current_spreadsheet, current_worksheet_name)

        def calculateCounts(self):
            self.setRangeForMBA()

    class Style:
        def __init__(self, cell):
            self.font = copy(cell.font)
            self.border = copy(cell.border)
            self.fill = copy(cell.fill)
            self.number_format = copy(cell.number_format)
            self.protection = copy(cell.protection)
            self.alignment = copy(cell.alignment)

        def copyToCell(self, new_cell):
            new_cell.font = copy(self.font)
            new_cell.border = copy(self.border)
            new_cell.fill = copy(self.fill)
            new_cell.number_format = copy(self.number_format)
            new_cell.protection = copy(self.protection)
            new_cell.alignment = copy(self.alignment)
